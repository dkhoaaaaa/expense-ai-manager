import os
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import text

from app import db


class AdminReportService:
    REPORTS = {
        "users": {
            "title": "BÁO CÁO NGƯỜI DÙNG",
            "sheetName": "Nguoi dung",
            "filename": "bao_cao_nguoi_dung",
            "dateColumn": "tk.ngay_tao",
            "columns": [
                {"key": "ID", "label": "ID", "type": "number"},
                {"key": "Email", "label": "Email", "type": "text"},
                {"key": "HoTen", "label": "Họ tên", "type": "text"},
                {"key": "SoDienThoai", "label": "Số điện thoại", "type": "text"},
                {"key": "VaiTro", "label": "Vai trò", "type": "text"},
                {"key": "TrangThai", "label": "Trạng thái", "type": "text"},
                {"key": "NgayTao", "label": "Ngày tạo", "type": "date"},
            ],
            "sql": """
                SELECT tk.id AS ID, tk.email AS Email, nd.ho_ten AS HoTen,
                       nd.so_dien_thoai AS SoDienThoai, tk.vai_tro AS VaiTro,
                       tk.trang_thai AS TrangThai, tk.ngay_tao AS NgayTao
                FROM tai_khoan tk
                LEFT JOIN nguoi_dung nd ON nd.tai_khoan_id = tk.id
            """,
            "orderBy": "tk.ngay_tao DESC, tk.id DESC",
        },
        "transactions": {
            "title": "BÁO CÁO GIAO DỊCH",
            "sheetName": "Giao dich",
            "filename": "bao_cao_giao_dich",
            "dateColumn": "gd.ngay_giao_dich",
            "columns": [
                {"key": "ID", "label": "ID", "type": "number"},
                {"key": "Email", "label": "Email", "type": "text"},
                {"key": "DanhMuc", "label": "Danh mục", "type": "text"},
                {"key": "Loai", "label": "Loại", "type": "text"},
                {"key": "SoTien", "label": "Số tiền", "type": "money"},
                {"key": "MoTa", "label": "Mô tả", "type": "text"},
                {"key": "NgayGiaoDich", "label": "Ngày giao dịch", "type": "date"},
                {"key": "PhuongThucPhanLoai", "label": "Phương thức", "type": "text"},
                {"key": "DoTinCay", "label": "Độ tin cậy", "type": "number"},
            ],
            "sql": """
                SELECT gd.id AS ID, tk.email AS Email, dm.ten_danh_muc AS DanhMuc,
                       gd.loai AS Loai, gd.so_tien AS SoTien, gd.mo_ta AS MoTa,
                       gd.ngay_giao_dich AS NgayGiaoDich,
                       gd.phuong_thuc_phan_loai AS PhuongThucPhanLoai,
                       gd.do_tin_cay AS DoTinCay
                FROM giao_dich gd
                LEFT JOIN tai_khoan tk ON tk.id = gd.tai_khoan_id
                LEFT JOIN danh_muc dm ON dm.id = gd.danh_muc_id
            """,
            "orderBy": "gd.ngay_giao_dich DESC, gd.id DESC",
        },
        "categories": {
            "title": "BÁO CÁO DANH MỤC",
            "sheetName": "Danh muc",
            "filename": "bao_cao_danh_muc",
            "dateColumn": "dm.ngay_tao",
            "columns": [
                {"key": "ID", "label": "ID", "type": "number"},
                {"key": "TenDanhMuc", "label": "Tên danh mục", "type": "text"},
                {"key": "Loai", "label": "Loại", "type": "text"},
                {"key": "KeywordAI", "label": "Keyword AI", "type": "text"},
                {"key": "TrangThai", "label": "Trạng thái", "type": "text"},
                {"key": "NgayTao", "label": "Ngày tạo", "type": "date"},
            ],
            "sql": """
                SELECT dm.id AS ID, dm.ten_danh_muc AS TenDanhMuc,
                       dm.loai AS Loai, dm.keyword_ai AS KeywordAI,
                       dm.trang_thai AS TrangThai, dm.ngay_tao AS NgayTao
                FROM danh_muc dm
            """,
            "orderBy": "dm.loai ASC, dm.ten_danh_muc ASC",
        },
        "premium": {
            "title": "BÁO CÁO PREMIUM",
            "sheetName": "Premium",
            "filename": "bao_cao_premium",
            "dateColumn": "gp.ngay_tao",
            "columns": [
                {"key": "ID", "label": "ID", "type": "number"},
                {"key": "Email", "label": "Email", "type": "text"},
                {"key": "TenGoi", "label": "Tên gói", "type": "text"},
                {"key": "Gia", "label": "Giá", "type": "money"},
                {"key": "TrangThai", "label": "Trạng thái", "type": "text"},
                {"key": "NgayBatDau", "label": "Ngày bắt đầu", "type": "date"},
                {"key": "NgayKetThuc", "label": "Ngày kết thúc", "type": "date"},
                {"key": "NgayTao", "label": "Ngày tạo", "type": "date"},
            ],
            "sql": """
                SELECT gp.id AS ID, tk.email AS Email, gp.ten_goi AS TenGoi,
                       gp.gia AS Gia, gp.trang_thai AS TrangThai,
                       gp.ngay_bat_dau AS NgayBatDau, gp.ngay_ket_thuc AS NgayKetThuc,
                       gp.ngay_tao AS NgayTao
                FROM goi_premium gp
                LEFT JOIN tai_khoan tk ON tk.id = gp.tai_khoan_id
            """,
            "orderBy": "gp.ngay_tao DESC, gp.id DESC",
        },
        "payments": {
            "title": "BÁO CÁO THANH TOÁN",
            "sheetName": "Thanh toan",
            "filename": "bao_cao_thanh_toan",
            "dateColumn": "COALESCE(tt.ngay_thanh_toan, tt.ngay_tao)",
            "columns": [
                {"key": "ID", "label": "ID", "type": "number"},
                {"key": "Email", "label": "Email", "type": "text"},
                {"key": "TenGoi", "label": "Tên gói", "type": "text"},
                {"key": "SoTien", "label": "Số tiền", "type": "money"},
                {"key": "PhuongThucThanhToan", "label": "Phương thức", "type": "text"},
                {"key": "TrangThaiThanhToan", "label": "Trạng thái", "type": "text"},
                {"key": "MaGiaoDich", "label": "Mã giao dịch", "type": "text"},
                {"key": "NgayThanhToan", "label": "Ngày thanh toán", "type": "date"},
                {"key": "NgayTao", "label": "Ngày tạo", "type": "date"},
            ],
            "sql": """
                SELECT tt.id AS ID, tk.email AS Email, gp.ten_goi AS TenGoi,
                       tt.so_tien AS SoTien, tt.phuong_thuc_thanh_toan AS PhuongThucThanhToan,
                       tt.trang_thai_thanh_toan AS TrangThaiThanhToan,
                       tt.ma_giao_dich AS MaGiaoDich, tt.ngay_thanh_toan AS NgayThanhToan,
                       tt.ngay_tao AS NgayTao
                FROM thanh_toan tt
                LEFT JOIN tai_khoan tk ON tk.id = tt.tai_khoan_id
                LEFT JOIN goi_premium gp ON gp.id = tt.goi_premium_id
            """,
            "orderBy": "COALESCE(tt.ngay_thanh_toan, tt.ngay_tao) DESC, tt.id DESC",
        },
        "ai": {
            "title": "BÁO CÁO AI MODEL",
            "sheetName": "AI Report",
            "filename": "bao_cao_ai_model",
            "dateColumn": "ls.ngay_tao",
            "columns": [
                {"key": "ID", "label": "ID", "type": "number"},
                {"key": "Email", "label": "Email", "type": "text"},
                {"key": "VanBanNhap", "label": "Văn bản nhập", "type": "text"},
                {"key": "DanhMucDuDoan", "label": "Danh mục dự đoán", "type": "text"},
                {"key": "DoTinCay", "label": "Độ tin cậy", "type": "number"},
                {"key": "TenModel", "label": "Tên model", "type": "text"},
                {"key": "NgayTao", "label": "Ngày tạo", "type": "date"},
            ],
            "sql": """
                SELECT ls.id AS ID, tk.email AS Email, ls.van_ban_nhap AS VanBanNhap,
                       dm.ten_danh_muc AS DanhMucDuDoan, ls.do_tin_cay AS DoTinCay,
                       ls.ten_model AS TenModel, ls.ngay_tao AS NgayTao
                FROM lich_su_ai_phan_loai ls
                LEFT JOIN tai_khoan tk ON tk.id = ls.tai_khoan_id
                LEFT JOIN danh_muc dm ON dm.id = ls.danh_muc_du_doan_id
            """,
            "orderBy": "ls.ngay_tao DESC, ls.id DESC",
        },
        "chatbot": {
            "title": "BÁO CÁO CHATBOT LOGS",
            "sheetName": "Chatbot",
            "filename": "bao_cao_chatbot",
            "dateColumn": "tn.ngay_tao",
            "columns": [
                {"key": "ID", "label": "ID", "type": "number"},
                {"key": "Email", "label": "Email", "type": "text"},
                {"key": "NguoiGui", "label": "Người gửi", "type": "text"},
                {"key": "NoiDung", "label": "Nội dung", "type": "text"},
                {"key": "NgayTao", "label": "Ngày tạo", "type": "date"},
            ],
            "sql": """
                SELECT tn.id AS ID, tk.email AS Email, tn.nguoi_gui AS NguoiGui,
                       tn.noi_dung AS NoiDung, tn.ngay_tao AS NgayTao
                FROM tin_nhan_chatbot tn
                LEFT JOIN tai_khoan tk ON tk.id = tn.tai_khoan_id
            """,
            "orderBy": "tn.ngay_tao DESC, tn.id DESC",
        },
    }

    @classmethod
    def buildReport(cls, reportType, fileFormat, fromDate=None, toDate=None):
        reportConfig = cls.REPORTS.get(reportType)
        if not reportConfig:
            raise ValueError("Loại báo cáo không hợp lệ")

        if fileFormat not in ("excel", "pdf"):
            raise ValueError("Định dạng báo cáo không hợp lệ")

        fromDateValue = cls._parseDate(fromDate, "Từ ngày") if fromDate else None
        toDateValue = cls._parseDate(toDate, "Đến ngày") if toDate else None

        if fromDateValue and toDateValue and fromDateValue > toDateValue:
            raise ValueError("Từ ngày không được lớn hơn đến ngày")

        rows = cls._fetchRows(reportConfig, fromDateValue, toDateValue)
        stats = cls._buildStats(reportType, rows)

        if fileFormat == "excel":
            fileBytes = cls._toExcel(reportConfig, rows, stats)
            extension = "xlsx"
            mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            fileBytes = cls._toPdf(reportConfig, rows, stats)
            extension = "pdf"
            mimetype = "application/pdf"

        filename = f"{reportConfig['filename']}.{extension}"
        return fileBytes, filename, mimetype

    @staticmethod
    def _parseDate(value, label):
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError as exc:
            raise ValueError(f"{label} không hợp lệ") from exc

    @staticmethod
    def _fetchRows(reportConfig, fromDate, toDate):
        conditions = []
        params = {}

        if fromDate:
            conditions.append(f"DATE({reportConfig['dateColumn']}) >= :fromDate")
            params["fromDate"] = fromDate

        if toDate:
            conditions.append(f"DATE({reportConfig['dateColumn']}) <= :toDate")
            params["toDate"] = toDate

        whereSql = f" WHERE {' AND '.join(conditions)}" if conditions else ""
        sql = f"{reportConfig['sql']} {whereSql} ORDER BY {reportConfig['orderBy']}"
        result = db.session.execute(text(sql), params)
        return [dict(row._mapping) for row in result]

    @staticmethod
    def _buildStats(reportType, rows):
        totalRows = len(rows)

        if reportType == "users":
            activeUsers = sum(1 for row in rows if row.get("TrangThai") == "ACTIVE")
            premiumUsers = sum(1 for row in rows if row.get("VaiTro") == "PREMIUM")
            return [
                ("Tổng người dùng", totalRows),
                ("Đang hoạt động", activeUsers),
                ("Người dùng Premium", premiumUsers),
            ]

        if reportType == "transactions":
            totalIncome = sum(AdminReportService._toNumber(row.get("SoTien")) for row in rows if row.get("Loai") == "THU")
            totalExpense = sum(AdminReportService._toNumber(row.get("SoTien")) for row in rows if row.get("Loai") == "CHI")
            return [
                ("Tổng giao dịch", totalRows),
                ("Tổng thu", AdminReportService._formatCurrency(totalIncome)),
                ("Tổng chi", AdminReportService._formatCurrency(totalExpense)),
            ]

        if reportType == "premium":
            activePackages = sum(1 for row in rows if row.get("TrangThai") == "ACTIVE")
            totalPackageValue = sum(AdminReportService._toNumber(row.get("Gia")) for row in rows)
            return [
                ("Tổng gói", totalRows),
                ("Gói đang hoạt động", activePackages),
                ("Tổng giá trị", AdminReportService._formatCurrency(totalPackageValue)),
            ]

        if reportType == "payments":
            successPayments = sum(1 for row in rows if row.get("TrangThaiThanhToan") == "SUCCESS")
            successAmount = sum(AdminReportService._toNumber(row.get("SoTien")) for row in rows if row.get("TrangThaiThanhToan") == "SUCCESS")
            return [
                ("Tổng thanh toán", totalRows),
                ("Thành công", successPayments),
                ("Doanh thu thành công", AdminReportService._formatCurrency(successAmount)),
            ]

        if reportType == "ai":
            confidenceValues = [
                AdminReportService._toNumber(row.get("DoTinCay"))
                for row in rows
                if row.get("DoTinCay") is not None
            ]
            averageConfidence = sum(confidenceValues) / len(confidenceValues) if confidenceValues else 0
            return [
                ("Tổng lượt AI phân loại", totalRows),
                ("Độ tin cậy trung bình", f"{averageConfidence:.2f}%"),
            ]

        userMessages = sum(1 for row in rows if row.get("NguoiGui") == "USER")
        botMessages = sum(1 for row in rows if row.get("NguoiGui") == "BOT")
        return [
            ("Tổng tin nhắn", totalRows),
            ("Tin nhắn USER", userMessages),
            ("Tin nhắn BOT", botMessages),
        ]

    @staticmethod
    def _toExcel(reportConfig, rows, stats):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = reportConfig["sheetName"]

        columns = reportConfig["columns"]
        lastColumn = get_column_letter(len(columns))
        titleFill = PatternFill("solid", fgColor="DDEBFF")
        headerFill = PatternFill("solid", fgColor="E8F1FF")
        statsFill = PatternFill("solid", fgColor="F8FAFC")
        borderSide = Side(style="thin", color="CBD5E1")
        thinBorder = Border(left=borderSide, right=borderSide, top=borderSide, bottom=borderSide)

        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        titleCell = worksheet["A1"]
        titleCell.value = reportConfig["title"]
        titleCell.font = Font(bold=True, size=16, color="1E3A8A")
        titleCell.fill = titleFill
        titleCell.alignment = Alignment(horizontal="center", vertical="center")
        worksheet.row_dimensions[1].height = 28

        worksheet["A2"] = f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        worksheet["A2"].font = Font(italic=True, color="475569")

        currentRow = 3
        for statLabel, statValue in stats:
            worksheet.cell(row=currentRow, column=1, value=statLabel)
            worksheet.cell(row=currentRow, column=2, value=statValue)
            worksheet.cell(row=currentRow, column=1).font = Font(bold=True)
            worksheet.cell(row=currentRow, column=1).fill = statsFill
            worksheet.cell(row=currentRow, column=2).fill = statsFill
            currentRow += 1

        headerRow = currentRow + 1
        for index, column in enumerate(columns, start=1):
            cell = worksheet.cell(row=headerRow, column=index, value=column["label"])
            cell.fill = headerFill
            cell.font = Font(bold=True, color="0F172A")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thinBorder

        if rows:
            for rowIndex, row in enumerate(rows, start=headerRow + 1):
                for columnIndex, column in enumerate(columns, start=1):
                    value = row.get(column["key"])
                    cell = worksheet.cell(row=rowIndex, column=columnIndex, value=value)
                    cell.border = thinBorder
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

                    if column["type"] == "money":
                        cell.value = AdminReportService._toNumber(value)
                        cell.number_format = '#,##0 "VNĐ"'
                    elif column["type"] == "date" and value:
                        cell.value = AdminReportService._toDateTime(value)
                        cell.number_format = "dd/mm/yyyy"
                    elif column["type"] == "number" and value is not None:
                        cell.value = AdminReportService._toNumber(value)
        else:
            emptyCell = worksheet.cell(row=headerRow + 1, column=1, value="Không có dữ liệu phù hợp")
            emptyCell.border = thinBorder
            worksheet.merge_cells(start_row=headerRow + 1, start_column=1, end_row=headerRow + 1, end_column=len(columns))

        worksheet.freeze_panes = f"A{headerRow + 1}"
        worksheet.auto_filter.ref = f"A{headerRow}:{lastColumn}{max(headerRow + len(rows), headerRow + 1)}"

        for columnIndex, column in enumerate(columns, start=1):
            columnLetter = get_column_letter(columnIndex)
            maxLength = len(column["label"])

            for cell in worksheet[columnLetter]:
                maxLength = max(maxLength, len(str(cell.value or "")))

            worksheet.column_dimensions[columnLetter].width = min(max(maxLength + 3, 12), 35)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    @staticmethod
    def _toPdf(reportConfig, rows, stats):
        output = BytesIO()
        document = SimpleDocTemplate(
            output,
            pagesize=landscape(A4),
            leftMargin=24,
            rightMargin=24,
            topMargin=28,
            bottomMargin=28,
        )

        styles = getSampleStyleSheet()
        fontName = AdminReportService._getPdfFontName()
        bodyStyle = ParagraphStyle("ReportBody", parent=styles["BodyText"], fontName=fontName, fontSize=8, leading=10)
        titleStyle = ParagraphStyle("ReportTitle", parent=styles["Title"], fontName=fontName, fontSize=16, textColor=colors.HexColor("#1E3A8A"))
        smallStyle = ParagraphStyle("ReportSmall", parent=styles["BodyText"], fontName=fontName, fontSize=9, textColor=colors.HexColor("#475569"))

        elements = [
            Paragraph(escape(reportConfig["title"]), titleStyle),
            Paragraph(f"Ngày xuất báo cáo: {datetime.now().strftime('%d/%m/%Y %H:%M')}", smallStyle),
            Spacer(1, 12),
        ]

        statsTable = Table(
            [[Paragraph(escape(str(label)), smallStyle), Paragraph(escape(str(value)), bodyStyle)] for label, value in stats],
            colWidths=[110, 120],
        )
        statsTable.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
                    ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
                    ("FONTNAME", (0, 0), (-1, -1), fontName),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ]
            )
        )
        elements.extend([statsTable, Spacer(1, 14)])

        columns = reportConfig["columns"]
        tableRows = [[Paragraph(escape(column["label"]), bodyStyle) for column in columns]]
        dataRows = rows or [{columns[0]["key"]: "Không có dữ liệu phù hợp"}]

        for row in dataRows:
            tableRows.append(
                [
                    Paragraph(escape(AdminReportService._formatDisplayValue(row.get(column["key"]), column["type"])), bodyStyle)
                    for column in columns
                ]
            )

        pageWidth = landscape(A4)[0] - document.leftMargin - document.rightMargin
        columnWidth = pageWidth / len(columns)
        dataTable = Table(tableRows, repeatRows=1, colWidths=[columnWidth] * len(columns))
        dataTable.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8F1FF")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CBD5E1")),
                    ("FONTNAME", (0, 0), (-1, 0), fontName),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        elements.append(dataTable)

        document.build(elements, onFirstPage=AdminReportService._drawPdfFooter, onLaterPages=AdminReportService._drawPdfFooter)
        output.seek(0)
        return output

    @staticmethod
    def _drawPdfFooter(canvas, document):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#64748B"))
        canvas.drawString(document.leftMargin, 15, "Expense AI Manager - Admin Report")
        canvas.drawRightString(landscape(A4)[0] - document.rightMargin, 15, f"Trang {document.page}")
        canvas.restoreState()

    @staticmethod
    def _formatDisplayValue(value, valueType):
        if value is None:
            return ""

        if valueType == "money":
            return AdminReportService._formatCurrency(value)

        if valueType == "date":
            dateValue = AdminReportService._toDateTime(value)
            return dateValue.strftime("%d/%m/%Y") if dateValue else ""

        return str(value)

    @staticmethod
    def _formatCurrency(value):
        return f"{AdminReportService._toNumber(value):,.0f} VNĐ"

    @staticmethod
    def _toNumber(value):
        if value is None or value == "":
            return 0

        if isinstance(value, Decimal):
            return float(value)

        return float(value)

    @staticmethod
    def _toDateTime(value):
        if isinstance(value, datetime):
            return value

        if isinstance(value, date):
            return datetime.combine(value, datetime.min.time())

        if isinstance(value, str):
            for dateFormat in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                try:
                    return datetime.strptime(value, dateFormat)
                except ValueError:
                    continue

        return None

    @staticmethod
    def _getPdfFontName():
        fontPaths = [
            r"C:\Windows\Fonts\arial.ttf",
            r"C:\Windows\Fonts\segoeui.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        ]

        for fontPath in fontPaths:
            if os.path.exists(fontPath):
                fontName = "ReportUnicode"
                if fontName not in pdfmetrics.getRegisteredFontNames():
                    pdfmetrics.registerFont(TTFont(fontName, fontPath))
                return fontName

        return "Helvetica"
